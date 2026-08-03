"""
Build a formula-driven CRE underwriting workbook from a JSON assumptions file.

    python3 build_model.py assumptions.json output.xlsx

Every projected cell is a live Excel formula pointing at a named assumption
range. Nothing that a reviewer would want to flex is baked in as a number, which
is what makes the workbook auditable rather than merely presentable.

Sheets: README | Assumptions | Revenue | Operating Expenses | Cash Flow |
        Debt Schedule | Returns | Sensitivity | Checks
"""

from __future__ import annotations

import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

import cre_engine as eng

# ---------------------------------------------------------------------------
# Presentation constants
# ---------------------------------------------------------------------------

INPUT_BLUE = Font(color="0000FF", name="Calibri", size=10)
FORMULA_BLACK = Font(color="000000", name="Calibri", size=10)
LINK_GREEN = Font(color="008000", name="Calibri", size=10)
LABEL = Font(color="000000", name="Calibri", size=10)
BOLD = Font(bold=True, name="Calibri", size=10)
TITLE = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
SECTION = Font(bold=True, size=10, color="FFFFFF", name="Calibri")

NAVY = PatternFill("solid", start_color="1F3864")
SLATE = PatternFill("solid", start_color="44546A")
LIGHT = PatternFill("solid", start_color="D9E1F2")
AMBER = PatternFill("solid", start_color="FFF2CC")

TOP_BORDER = Border(top=Side(style="thin"))
TOTAL_BORDER = Border(top=Side(style="thin"), bottom=Side(style="double"))

CUR = '$#,##0;($#,##0);"—"'
CUR2 = '$#,##0.00;($#,##0.00);"—"'
PCT = '0.0%;(0.0%);"—"'
PCT2 = '0.00%;(0.00%);"—"'
MULT = '0.00"x";;"—"'
NUM = '#,##0;(#,##0);"—"'


def col(year: int) -> str:
    """Year 0 -> B, year 1 -> C, ..."""
    return get_column_letter(2 + year)


# ---------------------------------------------------------------------------
# Cash Flow row map — every cross-sheet formula resolves through this
# ---------------------------------------------------------------------------

CF = {
    "gpr": 5,
    "recoveries": 6,
    "other": 7,
    "pgi": 8,
    "vacancy": 9,
    "egi": 10,
    "opex": 11,
    "mgmt": 12,
    "opex_total": 13,
    "noi": 14,
    "reserve": 16,
    "ti": 17,
    "lc": 18,
    "capital": 19,
    "cfbds": 20,
    "ds": 22,
    "cfads": 23,
    "dscr": 25,
    "dy_orig": 26,
    "dy_curr": 27,
    "ltv": 28,
    "fwd_noi": 30,
    "gross_exit": 31,
    "disp": 32,
    "net_sale": 33,
    "payoff": 34,
    "net_equity": 35,
    "unlev_op": 37,
    "unlev_total": 38,
    "lev_op": 39,
    "lev_total": 40,
}

DS = {"begin": 5, "interest": 6, "principal": 7, "service": 8, "end": 9}


# ---------------------------------------------------------------------------
# Builder
# ---------------------------------------------------------------------------


class ModelBuilder:
    def __init__(self, a: dict):
        self.a = a
        self.warnings = eng.validate(a)
        self.m = eng.compute(a)
        self.hold = self.m["hold_years"]
        self.n = self.hold + 1  # project one extra year for forward NOI
        self.ptype = a.get("deal", {}).get("property_type", "other")
        self.is_mf = self.ptype == "multifamily"
        self.has_debt = self.m["debt"]["loan_amount"] > 0
        self.wb = Workbook()
        self.names: dict[str, str] = {}

    # -- helpers ---------------------------------------------------------

    def name(self, key: str, sheet: str, cell: str) -> None:
        ref = f"'{sheet}'!${cell[0]}${cell[1:]}"
        self.wb.defined_names.add(DefinedName(key, attr_text=ref))
        self.names[key] = key

    def title(self, ws, text: str, width: int) -> None:
        ws["A1"] = text
        ws["A1"].font = TITLE
        for c in range(1, width + 1):
            ws.cell(row=1, column=c).fill = NAVY
        ws.row_dimensions[1].height = 24

    def section(self, ws, row: int, text: str, width: int) -> None:
        ws.cell(row=row, column=1, value=text).font = SECTION
        for c in range(1, width + 1):
            ws.cell(row=row, column=c).fill = SLATE

    def year_header(self, ws, row: int, start: int = 0, end: int | None = None) -> None:
        end = self.n if end is None else end
        for y in range(start, end + 1):
            label = "Year 0" if y == 0 else f"Year {y}"
            if y == self.hold:
                label += "\n(Exit)"
            if y == self.n:
                label += "\n(Forward)"
            c = ws[f"{col(y)}{row}"]
            c.value = label
            c.font = BOLD
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.fill = LIGHT
            c.border = TOP_BORDER
        ws.row_dimensions[row].height = 28

    def label(self, ws, row: int, text: str, bold: bool = False, indent: int = 0) -> None:
        c = ws.cell(row=row, column=1, value=text)
        c.font = BOLD if bold else LABEL
        if indent:
            c.alignment = Alignment(indent=indent)

    # -- sheets ----------------------------------------------------------

    def build(self) -> Workbook:
        self.wb.remove(self.wb.active)
        self.sheet_assumptions()
        self.sheet_revenue()
        self.sheet_opex()
        self.sheet_cashflow()
        self.sheet_debt()
        self.sheet_returns()
        self.sheet_sensitivity()
        self.sheet_checks()
        self.sheet_readme()
        self.wb.move_sheet("README", offset=-len(self.wb.sheetnames) + 1)
        return self.wb

    # ---------------- Assumptions ----------------

    def sheet_assumptions(self) -> None:
        a = self.a
        ws = self.wb.create_sheet("Assumptions")
        self.title(ws, f"ASSUMPTIONS — {a.get('deal', {}).get('name', 'Untitled Deal')}", 5)
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 52
        sources = a.get("sources", {})

        r = 3
        ws[f"A{r}"] = "Input"
        ws[f"B{r}"] = "Value"
        ws[f"C{r}"] = "Unit"
        ws[f"D{r}"] = "Source / basis  (fill this in — it is the audit trail)"
        for c in "ABCD":
            ws[f"{c}{r}"].font = BOLD
            ws[f"{c}{r}"].fill = LIGHT
        r += 1

        def put(key, label, value, unit, fmt, src_key=None):
            nonlocal r
            ws[f"A{r}"] = label
            cell = ws[f"B{r}"]
            cell.value = value
            cell.font = INPUT_BLUE
            cell.number_format = fmt
            ws[f"C{r}"] = unit
            ws[f"C{r}"].font = Font(size=9, color="808080")
            src = sources.get(src_key or key)
            ws[f"D{r}"] = src if src else "⚠ UNSOURCED — replace with comp / T-12 / term sheet"
            ws[f"D{r}"].font = Font(size=9, color="000000" if src else "C00000")
            if not src:
                ws[f"D{r}"].fill = AMBER
            self.name(key, "Assumptions", f"B{r}")
            r += 1

        def head(text):
            nonlocal r
            self.section(ws, r, text, 4)
            r += 1

        g = eng._get
        head("PROPERTY")
        if self.is_mf:
            put("Asmp_Units", "Units", g(a, "property", "units"), "units", NUM)
            put("Asmp_SF", "Rentable SF (optional)", g(a, "property", "rentable_sf", default=0), "SF", NUM)
        else:
            put("Asmp_SF", "Rentable SF", g(a, "property", "rentable_sf"), "SF", NUM)
            put("Asmp_Units", "Units (n/a)", 0, "units", NUM)

        head("ACQUISITION")
        put("Asmp_Price", "Purchase price", g(a, "acquisition", "purchase_price"), "$", CUR)
        put("Asmp_AcqCostPct", "Acquisition costs", g(a, "acquisition", "acquisition_cost_pct", default=0), "% of price", PCT2)
        put("Asmp_UpfrontCapex", "Upfront capital / reserves", g(a, "acquisition", "upfront_capex", default=0), "$", CUR)

        head("REVENUE")
        if self.is_mf:
            itemised = bool(g(a, "revenue", "unit_types"))
            put("Asmp_RentUnit",
                "Market rent / unit / month" + (" — MEMO only, mix drives GPR" if itemised else ""),
                g(a, "revenue", "market_rent_month", default=0), "$/unit/mo", CUR2)
            put("Asmp_RentPSF", "Market rent (n/a)", 0, "$/SF/yr", CUR2)
        else:
            put("Asmp_RentPSF", "Market rent", g(a, "revenue", "market_rent_psf", default=0), "$/SF/yr", CUR2)
            put("Asmp_RentUnit", "Market rent per unit (n/a)", 0, "$/unit/mo", CUR2)
        put("Asmp_RentGrowth", "Market rent growth", g(a, "revenue", "rent_growth", default=0), "%/yr", PCT2)
        put("Asmp_Vacancy", "Vacancy & credit loss", g(a, "revenue", "vacancy_credit_loss", default=0), "% of PGI", PCT2)
        put("Asmp_OtherInc", "Other income", g(a, "revenue", "other_income_annual", default=0), "$/yr", CUR)
        put("Asmp_OtherGrowth", "Other income growth", g(a, "revenue", "other_income_growth", default=g(a, "revenue", "rent_growth", default=0)), "%/yr", PCT2)
        put("Asmp_RecovRatio", "Expense recovery ratio", g(a, "revenue", "recovery_ratio", default=0), "% of recoverable", PCT2)

        head("OPERATING EXPENSES")
        put("Asmp_OpExRatio", "OpEx ratio (Year 1 basis)", g(a, "expenses", "opex_ratio", default=0), "% of Yr-1 EGI", PCT2)
        put("Asmp_MgmtFee", "Management fee", g(a, "expenses", "management_fee_pct", default=0), "% of EGI", PCT2)
        put("Asmp_ExpGrowth", "Expense growth", g(a, "expenses", "expense_growth", default=0), "%/yr", PCT2)

        head("CAPITAL COSTS")
        if self.is_mf:
            put("Asmp_ReserveUnit", "Replacement reserve", g(a, "expenses", "capex_reserve_per_unit", default=0), "$/unit/yr", CUR2)
            put("Asmp_ReservePSF", "Reserve (n/a)", 0, "$/SF/yr", CUR2)
            put("Asmp_TurnRate", "Annual turnover rate", g(a, "expenses", "turnover_rate", default=0), "% of units", PCT2)
            put("Asmp_TurnCost", "Unit turn cost", g(a, "expenses", "turn_cost_per_unit", default=0), "$/turn", CUR)
            for k in ("Asmp_RollPct", "Asmp_Renewal", "Asmp_TInew", "Asmp_TIrenew", "Asmp_LCnew", "Asmp_LCrenew", "Asmp_Term"):
                put(k, f"{k.replace('Asmp_', '')} (n/a for multifamily)", 0, "", NUM)
        else:
            put("Asmp_ReservePSF", "Replacement reserve", g(a, "expenses", "capex_reserve_psf", default=0), "$/SF/yr", CUR2)
            put("Asmp_ReserveUnit", "Reserve (n/a)", 0, "$/unit/yr", CUR2)
            put("Asmp_RollPct", "SF rolling per year", g(a, "expenses", "rollover_pct_annual", default=0), "% of SF", PCT2)
            put("Asmp_Renewal", "Renewal probability", g(a, "expenses", "renewal_probability", default=0.6), "%", PCT2)
            put("Asmp_TInew", "TI — new lease", g(a, "expenses", "ti_new_psf", default=0), "$/SF", CUR2)
            put("Asmp_TIrenew", "TI — renewal", g(a, "expenses", "ti_renewal_psf", default=0), "$/SF", CUR2)
            put("Asmp_LCnew", "LC — new lease", g(a, "expenses", "lc_new_pct", default=0), "% of lease value", PCT2)
            put("Asmp_LCrenew", "LC — renewal", g(a, "expenses", "lc_renewal_pct", default=0), "% of lease value", PCT2)
            put("Asmp_Term", "New lease term", g(a, "expenses", "new_lease_term_years", default=5), "years", NUM)
            put("Asmp_TurnRate", "Turnover (n/a)", 0, "", PCT2)
            put("Asmp_TurnCost", "Turn cost (n/a)", 0, "", CUR)

        head("DEBT")
        put("Asmp_Loan", "Loan amount", self.m["debt"]["loan_amount"], "$", CUR, src_key="loan_amount")
        put("Asmp_Rate", "Interest rate", g(a, "debt", "interest_rate", default=0), "%/yr", PCT2)
        put("Asmp_IO", "Interest-only period", g(a, "debt", "io_years", default=0), "years", NUM)
        put("Asmp_Amort", "Amortization term", g(a, "debt", "amort_years", default=30), "years", NUM)
        put("Asmp_DSCRMin", "DSCR covenant", g(a, "debt", "dscr_min", default=1.25), "x", MULT)
        put("Asmp_DYMin", "Debt yield covenant", g(a, "debt", "debt_yield_min", default=0.09), "%", PCT2)

        head("EXIT & RETURNS")
        put("Asmp_Hold", "Hold period", self.hold, "years", NUM)
        put("Asmp_ExitCap", "Exit cap rate", g(a, "exit", "exit_cap_rate"), "%", PCT2)
        put("Asmp_DispPct", "Disposition costs", g(a, "exit", "disposition_cost_pct", default=0), "% of gross", PCT2)
        put("Asmp_DiscRate", "Discount rate (unlevered)", g(a, "returns", "discount_rate", default=0), "%", PCT2)

        r += 1
        ws[f"A{r}"] = "Blue = input you can change.  Black = formula.  Green = link to another sheet."
        ws[f"A{r}"].font = Font(italic=True, size=9, color="808080")

    # ---------------- Revenue ----------------

    def sheet_revenue(self) -> None:
        """
        Three revenue modes. Whichever is used, the sheet ends with a single
        Gross Potential Rent row that Cash Flow links to, so the downstream
        model never has to care how revenue was built.
        """
        ws = self.wb.create_sheet("Revenue")
        self.title(ws, "REVENUE BUILD-UP", self.n + 2)
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 12
        for y in range(1, self.n + 1):
            ws.column_dimensions[col(y)].width = 15

        a = self.a
        mode = eng._get(a, "revenue", "mode", default="simple")
        unit_types = eng._get(a, "revenue", "unit_types", default=[])
        tenants = eng._get(a, "revenue", "tenants", default=[])
        rows: dict[str, int] = {}

        # ---- Itemised input block (unit mix / rent roll) --------------------
        item_rows: list[int] = []
        r = 3
        if mode == "unit_mix" and unit_types:
            self.section(ws, r, "UNIT MIX", 5)
            r += 1
            for i, h in enumerate(["Unit type", "Units", "SF/unit", "Rent $/mo", "Annual rent"]):
                c = ws.cell(row=r, column=1 + i, value=h)
                c.font = BOLD
                c.fill = LIGHT
            r += 1
            first = r
            for u in unit_types:
                ws.cell(row=r, column=1, value=u["name"])
                for j, (v, fmt) in enumerate(
                    [(u["count"], NUM), (u.get("sf", 0), NUM), (u["rent_month"], CUR2)]
                ):
                    c = ws.cell(row=r, column=2 + j, value=v)
                    c.font = INPUT_BLUE
                    c.number_format = fmt
                c = ws.cell(row=r, column=5, value=f"=B{r}*D{r}*12")
                c.number_format = CUR
                item_rows.append(r)
                r += 1
            last = r - 1
            ws.cell(row=r, column=1, value="Total").font = BOLD
            for cl in ("B", "E"):
                c = ws[f"{cl}{r}"]
                c.value = f"=SUM({cl}{first}:{cl}{last})"
                c.font = BOLD
                c.number_format = NUM if cl == "B" else CUR
                c.border = TOP_BORDER
            rows["mix_total_units"] = r
            r += 2

        elif mode == "rent_roll" and tenants:
            self.section(ws, r, "RENT ROLL — IN-PLACE LEASES", 5)
            r += 1
            for i, h in enumerate(
                ["Tenant", "SF", "Rent $/SF", "Lease end", "Annual rent", "Escalation"]
            ):
                c = ws.cell(row=r, column=1 + i, value=h)
                c.font = BOLD
                c.fill = LIGHT
            ws.column_dimensions["F"].width = 12
            r += 1
            first = r
            default_esc = eng._get(a, "revenue", "rent_growth", default=0.0)
            for t in tenants:
                ws.cell(row=r, column=1, value=t["name"])
                for j, (v, fmt) in enumerate(
                    [(t["sf"], NUM), (t["rent_psf"], CUR2), (t.get("lease_end_year", ""), "General")]
                ):
                    c = ws.cell(row=r, column=2 + j, value=v)
                    c.font = INPUT_BLUE
                    c.number_format = fmt
                c = ws.cell(row=r, column=5, value=f"=B{r}*C{r}")
                c.number_format = CUR
                # Escalation lives in its own input cell, not inside the projection
                # formula, so a reviewer can see and flex every lease's bump.
                e = ws.cell(row=r, column=6, value=t.get("escalation", default_esc))
                e.font = INPUT_BLUE
                e.number_format = PCT2
                item_rows.append(r)
                r += 1
            last = r - 1
            ws.cell(row=r, column=1, value="Total in-place").font = BOLD
            for cl in ("B", "E"):
                c = ws[f"{cl}{r}"]
                c.value = f"=SUM({cl}{first}:{cl}{last})"
                c.font = BOLD
                c.number_format = NUM if cl == "B" else CUR
                c.border = TOP_BORDER
            rows["roll_total"] = r
            r += 2

        # ---- Projection ----------------------------------------------------
        self.section(ws, r, "GROSS POTENTIAL RENT PROJECTION", self.n + 2)
        r += 1
        self.year_header(ws, r, start=1)
        r += 1

        if mode == "unit_mix" and unit_types:
            first_proj = r
            for i, u in enumerate(unit_types):
                self.label(ws, r, f"  {u['name']}")
                src = item_rows[i]
                for y in range(1, self.n + 1):
                    c = ws[f"{col(y)}{r}"]
                    c.value = f"=$E${src}*(1+Asmp_RentGrowth)^{y - 1}"
                    c.number_format = CUR
                r += 1
            proj_last = r - 1
            gpr_formula = lambda y: f"=SUM({col(y)}{first_proj}:{col(y)}{proj_last})"

        elif mode == "rent_roll" and tenants:
            first_proj = r
            for i, t in enumerate(tenants):
                self.label(ws, r, f"  {t['name']}")
                src = item_rows[i]
                for y in range(1, self.n + 1):
                    c = ws[f"{col(y)}{r}"]
                    c.value = f"=$E${src}*(1+$F${src})^{y - 1}"
                    c.number_format = CUR
                r += 1
            proj_last = r - 1
            gpr_formula = lambda y: f"=SUM({col(y)}{first_proj}:{col(y)}{proj_last})"

        elif self.is_mf:
            # Multifamily without an itemised mix: units x blended monthly rent.
            self.label(ws, r, "Units")
            for y in range(1, self.n + 1):
                ws[f"{col(y)}{r}"] = "=Asmp_Units"
                ws[f"{col(y)}{r}"].number_format = NUM
            u_row = r
            r += 1
            self.label(ws, r, "Market rent  ($/unit/month)")
            for y in range(1, self.n + 1):
                ws[f"{col(y)}{r}"] = f"=Asmp_RentUnit*(1+Asmp_RentGrowth)^{y - 1}"
                ws[f"{col(y)}{r}"].number_format = CUR2
            mr_row = r
            r += 1
            gpr_formula = lambda y: f"={col(y)}{u_row}*{col(y)}{mr_row}*12"

        else:
            self.label(ws, r, "Rentable SF")
            for y in range(1, self.n + 1):
                ws[f"{col(y)}{r}"] = "=Asmp_SF"
                ws[f"{col(y)}{r}"].number_format = NUM
            sf_row = r
            r += 1
            self.label(ws, r, "Market rent  ($/SF/year)")
            for y in range(1, self.n + 1):
                ws[f"{col(y)}{r}"] = f"=Asmp_RentPSF*(1+Asmp_RentGrowth)^{y - 1}"
                ws[f"{col(y)}{r}"].number_format = CUR2
            rent_row = r
            r += 1
            gpr_formula = lambda y: f"={col(y)}{sf_row}*{col(y)}{rent_row}"

        self.label(ws, r, "GROSS POTENTIAL RENT", bold=True)
        for y in range(1, self.n + 1):
            c = ws[f"{col(y)}{r}"]
            c.value = gpr_formula(y)
            c.number_format = CUR
            c.font = BOLD
            c.border = TOP_BORDER
        rows["gpr"] = r
        self.rev_rows = rows

        r += 2
        if mode == "rent_roll" and tenants:
            note = ("Each lease is grown at its own escalation. This template does NOT model lease "
                    "expiry, downtime, or mark-to-market at rollover — see references/rent-roll-modeling.md "
                    "for how to extend it when WALT and rollover risk drive the deal.")
        else:
            note = "Rent is grown at Asmp_RentGrowth each year. Vacancy is applied on the Cash Flow tab, not here."
        self.label(ws, r, note)
        ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="808080")

    # ---------------- Operating Expenses ----------------

    def sheet_opex(self) -> None:
        ws = self.wb.create_sheet("Operating Expenses")
        self.title(ws, "OPERATING EXPENSES", self.n + 2)
        ws.column_dimensions["A"].width = 38
        for y in range(1, self.n + 1):
            ws.column_dimensions[col(y)].width = 15
        self.year_header(ws, 3, start=1)

        mode = self.a.get("expenses", {}).get("mode", "ratio")
        line_items = self.a.get("expenses", {}).get("line_items", [])
        r = 4
        self.opex_rows = {}

        if mode == "line_items" and line_items:
            # Column B holds each line item's own growth rate as a visible input.
            # Baking the rate into the formula (=C4*(1+0.035)) would hide it from
            # review and make it impossible to flex — the exact defect this skill
            # teaches people to look for.
            hdr = ws.cell(row=r - 1, column=2, value="Growth")
            hdr.font = BOLD
            hdr.fill = LIGHT
            hdr.alignment = Alignment(horizontal="center")
            ws.column_dimensions["B"].width = 10

            first = r
            recoverable_rows = []
            for li in line_items:
                self.label(ws, r, ("  ↳ " if li.get("recoverable") else "     ") + li["name"])
                gcell = ws.cell(
                    row=r, column=2,
                    value=li.get("growth", self.a["expenses"].get("expense_growth", 0)),
                )
                gcell.font = INPUT_BLUE
                gcell.number_format = PCT2
                for y in range(1, self.n + 1):
                    c = ws[f"{col(y)}{r}"]
                    if y == 1:
                        c.value = li["amount_annual"]
                        c.font = INPUT_BLUE
                    else:
                        c.value = f"={col(y - 1)}{r}*(1+$B{r})"
                    c.number_format = CUR
                if li.get("recoverable"):
                    recoverable_rows.append(r)
                r += 1
            last = r - 1
            self.label(ws, r, "Total operating expenses (before mgmt fee)", bold=True)
            for y in range(1, self.n + 1):
                c = ws[f"{col(y)}{r}"]
                c.value = f"=SUM({col(y)}{first}:{col(y)}{last})"
                c.number_format = CUR
                c.font = BOLD
                c.border = TOP_BORDER
            self.opex_rows["base"] = r
            r += 1
            self.label(ws, r, "Recoverable pool  (↳ items above)")
            for y in range(1, self.n + 1):
                c = ws[f"{col(y)}{r}"]
                c.value = "=" + "+".join(f"{col(y)}{rr}" for rr in recoverable_rows) if recoverable_rows else 0
                c.number_format = CUR
            self.opex_rows["recoverable"] = r
        else:
            self.label(ws, r, "Operating expenses  (Yr-1 ratio, inflated)", bold=True)
            for y in range(1, self.n + 1):
                c = ws[f"{col(y)}{r}"]
                c.value = f"='Cash Flow'!C{CF['egi']}*Asmp_OpExRatio*(1+Asmp_ExpGrowth)^{y - 1}"
                c.number_format = CUR
                c.font = LINK_GREEN
            self.opex_rows["base"] = r
            r += 1
            self.label(ws, r, "Recoverable pool  (none in ratio mode)")
            for y in range(1, self.n + 1):
                ws[f"{col(y)}{r}"] = 0
                ws[f"{col(y)}{r}"].number_format = CUR
            self.opex_rows["recoverable"] = r

        r += 2
        note = (
            "Ratio mode anchors OpEx to YEAR-1 EGI and then inflates at Asmp_ExpGrowth. Applying the ratio "
            "to each year's EGI would force expenses to grow at the revenue rate and hide margin change."
            if mode != "line_items"
            else "Recoverable items (↳) feed the recovery line on Cash Flow at Asmp_RecovRatio."
        )
        self.label(ws, r, note)
        ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="808080")

    # ---------------- Cash Flow ----------------

    def sheet_cashflow(self) -> None:
        ws = self.wb.create_sheet("Cash Flow")
        self.title(ws, "ANNUAL CASH FLOW PRO FORMA", self.n + 2)
        ws.column_dimensions["A"].width = 42
        for y in range(0, self.n + 1):
            ws.column_dimensions[col(y)].width = 15
        self.year_header(ws, 4, start=0)
        R, gr = CF, self.rev_rows["gpr"]

        def row(key, label, formula_fn, fmt=CUR, bold=False, font=None, years=None):
            r = R[key]
            self.label(ws, r, label, bold=bold)
            for y in years if years is not None else range(1, self.n + 1):
                f = formula_fn(y)
                if f is None:
                    continue
                c = ws[f"{col(y)}{r}"]
                c.value = f
                c.number_format = fmt
                c.font = font or (BOLD if bold else FORMULA_BLACK)

        row("gpr", "Gross Potential Rent", lambda y: f"=Revenue!{col(y)}{gr}", font=LINK_GREEN)
        row("recoveries", "Expense Recoveries",
            lambda y: f"='Operating Expenses'!{col(y)}{self.opex_rows['recoverable']}*Asmp_RecovRatio", font=LINK_GREEN)
        row("other", "Other Income", lambda y: f"=Asmp_OtherInc*(1+Asmp_OtherGrowth)^{y - 1}")
        row("pgi", "Potential Gross Income",
            lambda y: f"=SUM({col(y)}{R['gpr']}:{col(y)}{R['other']})", bold=True)
        row("vacancy", "Less: Vacancy & Credit Loss", lambda y: f"=-{col(y)}{R['pgi']}*Asmp_Vacancy")
        row("egi", "EFFECTIVE GROSS INCOME",
            lambda y: f"={col(y)}{R['pgi']}+{col(y)}{R['vacancy']}", bold=True)
        row("opex", "Operating Expenses",
            lambda y: f"=-'Operating Expenses'!{col(y)}{self.opex_rows['base']}", font=LINK_GREEN)
        row("mgmt", "Management Fee", lambda y: f"=-{col(y)}{R['egi']}*Asmp_MgmtFee")
        row("opex_total", "Total Operating Expenses",
            lambda y: f"={col(y)}{R['opex']}+{col(y)}{R['mgmt']}", bold=True)
        row("noi", "NET OPERATING INCOME",
            lambda y: f"={col(y)}{R['egi']}+{col(y)}{R['opex_total']}", bold=True)
        for y in range(1, self.n + 1):
            ws[f"{col(y)}{R['noi']}"].border = TOP_BORDER

        self.section(ws, R["reserve"] - 1, "CAPITAL COSTS", self.n + 2)
        if self.is_mf:
            row("reserve", "Replacement Reserve",
                lambda y: f"=-Asmp_Units*Asmp_ReserveUnit*(1+Asmp_ExpGrowth)^{y - 1}")
            row("ti", "Unit Turn Cost",
                lambda y: f"=-Asmp_Units*Asmp_TurnRate*Asmp_TurnCost*(1+Asmp_ExpGrowth)^{y - 1}")
            row("lc", "Leasing Commissions (n/a)", lambda y: "=0")
        else:
            row("reserve", "Replacement Reserve",
                lambda y: f"=-Asmp_SF*Asmp_ReservePSF*(1+Asmp_ExpGrowth)^{y - 1}")
            row("ti", "Tenant Improvements",
                lambda y: "=-Asmp_SF*Asmp_RollPct*(Asmp_Renewal*Asmp_TIrenew+(1-Asmp_Renewal)*Asmp_TInew)")
            row("lc", "Leasing Commissions",
                lambda y: f"=-Asmp_SF*Asmp_RollPct*Asmp_RentPSF*(1+Asmp_RentGrowth)^{y - 1}"
                          "*Asmp_Term*(Asmp_Renewal*Asmp_LCrenew+(1-Asmp_Renewal)*Asmp_LCnew)")
        row("capital", "Total Capital Costs",
            lambda y: f"=SUM({col(y)}{R['reserve']}:{col(y)}{R['lc']})", bold=True)
        row("cfbds", "CASH FLOW BEFORE DEBT SERVICE",
            lambda y: f"={col(y)}{R['noi']}+{col(y)}{R['capital']}", bold=True)
        for y in range(1, self.n + 1):
            ws[f"{col(y)}{R['cfbds']}"].border = TOP_BORDER

        self.section(ws, R["ds"] - 1, "DEBT", self.n + 2)
        row("ds", "Debt Service",
            lambda y: f"=-'Debt Schedule'!{col(y)}{DS['service']}", font=LINK_GREEN)
        row("cfads", "CASH FLOW AFTER DEBT SERVICE",
            lambda y: f"={col(y)}{R['cfbds']}+{col(y)}{R['ds']}", bold=True)
        for y in range(1, self.n + 1):
            ws[f"{col(y)}{R['cfads']}"].border = TOP_BORDER

        self.section(ws, R["dscr"] - 1, "COVERAGE & LEVERAGE TESTS", self.n + 2)
        row("dscr", "DSCR  (NOI / Debt Service)",
            lambda y: f'=IF(-{col(y)}{R["ds"]}>0,{col(y)}{R["noi"]}/-{col(y)}{R["ds"]},"—")', fmt=MULT)
        row("dy_orig", "Debt Yield — origination  (NOI / original loan)",
            lambda y: f'=IF(Asmp_Loan>0,{col(y)}{R["noi"]}/Asmp_Loan,"—")', fmt=PCT2)
        row("dy_curr", "Debt Yield — current balance  (surveillance only)",
            lambda y: f"=IF('Debt Schedule'!{col(y)}{DS['begin']}>0,"
                      f'{col(y)}{R["noi"]}/\'Debt Schedule\'!{col(y)}{DS["begin"]},"—")', fmt=PCT2)
        row("ltv", "LTV on cost  (ending balance / price)",
            lambda y: f'=IF(Asmp_Price>0,\'Debt Schedule\'!{col(y)}{DS["end"]}/Asmp_Price,"—")', fmt=PCT)

        # ---- Exit block: single column at the hold year ----
        self.section(ws, R["fwd_noi"] - 1, "EXIT / REVERSION", self.n + 2)
        h = col(self.hold)
        fwd = col(self.n)
        exit_rows = [
            ("fwd_noi", f"Forward NOI  (Year {self.n})", f"={fwd}{R['noi']}", CUR),
            ("gross_exit", "Gross Exit Value  (Forward NOI / Exit Cap)",
             f'=IF(Asmp_ExitCap>0,{h}{R["fwd_noi"]}/Asmp_ExitCap,"—")', CUR),
            ("disp", "Less: Disposition Costs", f"=-{h}{R['gross_exit']}*Asmp_DispPct", CUR),
            ("net_sale", "Net Sale Proceeds", f"={h}{R['gross_exit']}+{h}{R['disp']}", CUR),
            ("payoff", "Less: Loan Payoff", f"=-'Debt Schedule'!{h}{DS['end']}", CUR),
            ("net_equity", "NET EQUITY PROCEEDS", f"={h}{R['net_sale']}+{h}{R['payoff']}", CUR),
        ]
        for key, lab, formula, fmt in exit_rows:
            r = R[key]
            self.label(ws, r, lab, bold=key == "net_equity")
            c = ws[f"{h}{r}"]
            c.value = formula
            c.number_format = fmt
            c.font = BOLD if key == "net_equity" else FORMULA_BLACK

        # ---- Return cash flow rows ----
        self.section(ws, R["unlev_op"] - 1, "CASH FLOWS FOR RETURN CALCULATIONS", self.n + 2)
        basis = "-(Asmp_Price+Asmp_Price*Asmp_AcqCostPct+Asmp_UpfrontCapex)"
        cfrows = [
            ("unlev_op", "Unlevered CF — operating only", basis,
             lambda y: f"={col(y)}{R['cfbds']}"),
            ("unlev_total", "Unlevered CF — total (incl. sale)", basis,
             lambda y: f"={col(y)}{R['cfbds']}" + (f"+{h}{R['net_sale']}" if y == self.hold else "")),
            ("lev_op", "Levered CF — operating only", f"{basis}+Asmp_Loan",
             lambda y: f"={col(y)}{R['cfads']}"),
            ("lev_total", "Levered CF — total (incl. reversion)", f"{basis}+Asmp_Loan",
             lambda y: f"={col(y)}{R['cfads']}" + (f"+{h}{R['net_equity']}" if y == self.hold else "")),
        ]
        for key, lab, y0, fn in cfrows:
            r = R[key]
            self.label(ws, r, lab, bold=True)
            c0 = ws[f"{col(0)}{r}"]
            c0.value = "=" + y0
            c0.number_format = CUR
            c0.font = BOLD
            for y in range(1, self.hold + 1):
                c = ws[f"{col(y)}{r}"]
                c.value = fn(y)
                c.number_format = CUR
                c.font = BOLD

        r = R["lev_total"] + 2
        self.label(ws, r, f"Year {self.n} is a FORWARD year: it exists only to supply exit NOI and is "
                          "excluded from every return calculation.")
        ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="808080")

    # ---------------- Debt Schedule ----------------

    def sheet_debt(self) -> None:
        ws = self.wb.create_sheet("Debt Schedule")
        self.title(ws, "DEBT SCHEDULE", self.n + 2)
        ws.column_dimensions["A"].width = 42
        for y in range(1, self.n + 1):
            ws.column_dimensions[col(y)].width = 15
        self.year_header(ws, 3, start=1)
        D = DS

        # Live IF() structure keeps the schedule correct when IO years change.
        # CUMIPMT/CUMPRINC always take the ORIGINAL principal, never the balance.
        period_start = "MAX(1,({y}-Asmp_IO-1)*12+1)"
        period_end = "MAX(1,({y}-Asmp_IO)*12)"

        specs = [
            ("begin", "Beginning Balance",
             lambda y: "=Asmp_Loan" if y == 1 else f"={col(y - 1)}{D['end']}"),
            ("interest", "Interest",
             lambda y: f"=IF(Asmp_Loan<=0,0,IF({y}<=Asmp_IO,{col(y)}{D['begin']}*Asmp_Rate,"
                       f"IF(Asmp_Rate=0,0,-CUMIPMT(Asmp_Rate/12,Asmp_Amort*12,Asmp_Loan,"
                       f"{period_start.format(y=y)},{period_end.format(y=y)},0))))"),
            ("principal", "Principal",
             lambda y: f"=IF(Asmp_Loan<=0,0,IF({y}<=Asmp_IO,0,"
                       f"IF(Asmp_Rate=0,Asmp_Loan/Asmp_Amort,"
                       f"-CUMPRINC(Asmp_Rate/12,Asmp_Amort*12,Asmp_Loan,"
                       f"{period_start.format(y=y)},{period_end.format(y=y)},0))))"),
            ("service", "Total Debt Service",
             lambda y: f"={col(y)}{D['interest']}+{col(y)}{D['principal']}"),
            ("end", "Ending Balance",
             lambda y: f"=MAX(0,{col(y)}{D['begin']}-{col(y)}{D['principal']})"),
        ]
        for key, lab, fn in specs:
            r = D[key]
            self.label(ws, r, lab, bold=key in ("service", "end"))
            for y in range(1, self.n + 1):
                c = ws[f"{col(y)}{r}"]
                c.value = fn(y)
                c.number_format = CUR
                c.font = BOLD if key in ("service", "end") else FORMULA_BLACK
            if key == "service":
                for y in range(1, self.n + 1):
                    ws[f"{col(y)}{r}"].border = TOP_BORDER

        r = D["end"] + 2
        self.label(ws, r, "Memo — annual constant (amortizing)")
        ws[f"C{r}"] = "=IF(Asmp_Loan>0,-PMT(Asmp_Rate/12,Asmp_Amort*12,Asmp_Loan)*12,0)"
        ws[f"C{r}"].number_format = CUR
        r += 1
        self.label(ws, r, "CUMIPMT / CUMPRINC take the ORIGINAL loan amount, never the current balance. "
                          "Using the balance understates interest in every year after the first.")
        ws.cell(row=r, column=1).font = Font(italic=True, size=9, color="808080")

    # ---------------- Returns ----------------

    def sheet_returns(self) -> None:
        ws = self.wb.create_sheet("Returns")
        self.title(ws, "RETURNS SUMMARY", 4)
        ws.column_dimensions["A"].width = 44
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 50
        R, h = CF, col(self.hold)
        last = col(self.hold)

        r = 3
        def put(label, formula, fmt, bold=False, note="", key=None):
            nonlocal r
            self.label(ws, r, label, bold=bold)
            c = ws[f"B{r}"]
            c.value = formula
            c.number_format = fmt
            c.font = BOLD if bold else FORMULA_BLACK
            if note:
                ws[f"C{r}"] = note
                ws[f"C{r}"].font = Font(size=9, color="808080")
            if key:
                self.name(key, "Returns", f"B{r}")
            r += 1

        def head(text):
            nonlocal r
            self.section(ws, r, text, 3)
            r += 1

        head("SOURCES & USES")
        put("Purchase price", "=Asmp_Price", CUR)
        put("Acquisition costs", "=Asmp_Price*Asmp_AcqCostPct", CUR)
        put("Upfront capital", "=Asmp_UpfrontCapex", CUR)
        put("Total basis (uses)", "=B4+B5+B6", CUR, bold=True, key="Ret_Basis")
        put("Loan proceeds", "=Asmp_Loan", CUR)
        put("Equity required", "=Ret_Basis-Asmp_Loan", CUR, bold=True, key="Ret_Equity")
        put("Sources − Uses  (must be zero)", "=(Asmp_Loan+Ret_Equity)-Ret_Basis", CUR,
            note="Balance check")

        head("GOING-IN METRICS")
        put("Year 1 NOI", f"='Cash Flow'!C{R['noi']}", CUR, key="Ret_NOI1")
        put("Going-in cap rate", "=IF(Asmp_Price>0,Ret_NOI1/Asmp_Price,0)", PCT2,
            note="Year 1 NOI ÷ purchase price")
        put("Price per SF / unit",
            "=IF(Asmp_Units>0,Asmp_Price/Asmp_Units,IF(Asmp_SF>0,Asmp_Price/Asmp_SF,0))", CUR2)

        head("RETURN METRICS")
        put("Unlevered IRR", f"=IRR('Cash Flow'!B{R['unlev_total']}:{last}{R['unlev_total']},0.1)",
            PCT2, bold=True, note="On total basis, incl. net sale proceeds")
        put("Levered IRR",
            f'=IF(Asmp_Loan>0,IRR(\'Cash Flow\'!B{R["lev_total"]}:{last}{R["lev_total"]},0.12),'
            f"IRR('Cash Flow'!B{R['unlev_total']}:{last}{R['unlev_total']},0.1))",
            PCT2, bold=True, note="On equity, incl. net equity proceeds")
        put("Equity multiple",
            f"=IF(Ret_Equity>0,(SUM('Cash Flow'!C{R['lev_op']}:{last}{R['lev_op']})"
            f"+'Cash Flow'!{h}{R['net_equity']})/Ret_Equity,0)",
            MULT, bold=True, note="Operating CF (no exit) + net equity proceeds, ÷ equity")
        put("Year 1 cash-on-cash",
            f"=IF(Ret_Equity>0,'Cash Flow'!C{R['lev_op']}/Ret_Equity,0)", PCT2)
        put("Average cash-on-cash",
            f"=IF(Ret_Equity>0,AVERAGE('Cash Flow'!C{R['lev_op']}:{last}{R['lev_op']})/Ret_Equity,0)",
            PCT2, note="Operating CF only — excludes the reversion")

        head("VALUE INDICATION")
        put("PV of operating cash flows",
            f"=NPV(Asmp_DiscRate,'Cash Flow'!C{R['unlev_op']}:{last}{R['unlev_op']})", CUR)
        put("PV of net sale proceeds",
            f"='Cash Flow'!{h}{R['net_sale']}/(1+Asmp_DiscRate)^Asmp_Hold", CUR)
        put("DCF indicated value", f"=B{r - 2}+B{r - 1}", CUR, bold=True, key="Ret_DCFValue")
        put("NPV vs. total basis", "=Ret_DCFValue-Ret_Basis", CUR, bold=True,
            note="Positive = priced below the DCF indication at this discount rate")

        head("COVENANT TESTS")
        dscr_rng = f"'Cash Flow'!C{R['dscr']}:{last}{R['dscr']}"
        dy_rng = f"'Cash Flow'!C{R['dy_orig']}:{last}{R['dy_orig']}"
        put("Minimum DSCR over hold", f'=IF(Asmp_Loan>0,MIN({dscr_rng}),"—")', MULT)
        put("DSCR covenant", "=Asmp_DSCRMin", MULT)
        put("DSCR test", f'=IF(Asmp_Loan<=0,"n/a — all cash",IF(MIN({dscr_rng})>=Asmp_DSCRMin,"PASS","FAIL"))', "General")
        io = self.a.get("debt", {}).get("io_years", 0)
        first_amort_col = col(io + 1) if io + 1 <= self.hold else None
        put("First amortizing-year DSCR",
            f"='Cash Flow'!{first_amort_col}{R['dscr']}" if first_amort_col else '="entire hold is IO"',
            MULT, note="The binding test — IO years flatter coverage")
        put("Minimum debt yield (origination)", f'=IF(Asmp_Loan>0,MIN({dy_rng}),"—")', PCT2)
        put("Debt yield covenant", "=Asmp_DYMin", PCT2)
        put("Debt yield test", f'=IF(Asmp_Loan<=0,"n/a — all cash",IF(MIN({dy_rng})>=Asmp_DYMin,"PASS","FAIL"))', "General")

        for rr in range(3, r):
            v = ws[f"B{rr}"].value
            if isinstance(v, str) and '"PASS"' in v:
                ws[f"B{rr}"].alignment = Alignment(horizontal="center")

    # ---------------- Sensitivity ----------------

    def sheet_sensitivity(self) -> None:
        ws = self.wb.create_sheet("Sensitivity")
        self.title(ws, "SENSITIVITY ANALYSIS", 9)
        ws.column_dimensions["A"].width = 30
        for c in range(2, 10):
            ws.column_dimensions[get_column_letter(c)].width = 13

        a, m = self.a, self.m
        base_cap = eng._get(a, "exit", "exit_cap_rate")
        base_growth = eng._get(a, "revenue", "rent_growth", default=0.0)
        base_rate = eng._get(a, "debt", "interest_rate", default=0.0)
        price = eng._get(a, "acquisition", "purchase_price")

        caps = [base_cap + d for d in (-0.0100, -0.0050, -0.0025, 0, 0.0025, 0.0050, 0.0100)]
        growths = [max(0.0, base_growth + d) for d in (-0.0150, -0.0100, -0.0050, 0, 0.0050, 0.0100, 0.0150)]

        # --- Table 1: Unlevered IRR (values from the engine; IRR needs a full re-projection)
        r = 3
        self.section(ws, r, "TABLE 1 — Unlevered IRR:  Exit Cap Rate (rows) × Rent Growth (cols)", 9)
        r += 1
        ws.cell(row=r, column=1, value="Exit Cap ↓ / Growth →").font = BOLD
        for j, gv in enumerate(growths):
            c = ws.cell(row=r, column=2 + j, value=gv)
            c.number_format = PCT2
            c.font = BOLD
            c.fill = LIGHT
        r += 1
        t1_top = r
        for cap in caps:
            c = ws.cell(row=r, column=1, value=cap)
            c.number_format = PCT2
            c.font = BOLD
            c.fill = LIGHT
            for j, gv in enumerate(growths):
                scen = json.loads(json.dumps(a))
                scen["exit"]["exit_cap_rate"] = cap
                scen["revenue"]["rent_growth"] = gv
                try:
                    val = eng.compute(scen)["returns"]["unlevered_irr"]
                except Exception:
                    val = None
                cc = ws.cell(row=r, column=2 + j, value=val)
                cc.number_format = PCT2
                if cap == base_cap and abs(gv - base_growth) < 1e-9:
                    cc.fill = AMBER
                    cc.font = BOLD
            r += 1
        r += 1
        ws.cell(row=r, column=1,
                value="Table 1 holds VALUES computed by the builder, because IRR requires re-running the whole "
                      "projection. Re-run build_model.py after changing assumptions, or convert it to a live "
                      "Excel Data Table (see README).").font = Font(italic=True, size=9, color="C00000")
        r += 2

        # --- Table 2: Year-1 DSCR — fully LIVE (single-formula metric)
        self.section(ws, r, "TABLE 2 — Year 1 DSCR:  Interest Rate (rows) × LTV (cols)   [LIVE FORMULAS]", 9)
        r += 1
        ltvs = [0.50, 0.55, 0.60, 0.65, 0.70, 0.75, 0.80]
        rates = [max(0.0001, base_rate + d) for d in (-0.0150, -0.0100, -0.0050, 0, 0.0050, 0.0100, 0.0150)]
        ws.cell(row=r, column=1, value="Rate ↓ / LTV →").font = BOLD
        for j, lv in enumerate(ltvs):
            c = ws.cell(row=r, column=2 + j, value=lv)
            c.number_format = PCT
            c.font = BOLD
            c.fill = LIGHT
        r += 1
        for rate in rates:
            c = ws.cell(row=r, column=1, value=rate)
            c.number_format = PCT2
            c.font = BOLD
            c.fill = LIGHT
            rl = f"$A{r}"
            for j, lv in enumerate(ltvs):
                cl = f"{get_column_letter(2 + j)}${r - 1}"
                # IO year 1 -> interest only; otherwise the amortizing constant.
                ds = (f"IF(Asmp_IO>=1,Asmp_Price*{cl}*{rl},"
                      f"-PMT({rl}/12,Asmp_Amort*12,Asmp_Price*{cl})*12)")
                cc = ws.cell(row=r, column=2 + j, value=f"=IF({ds}>0,Ret_NOI1/({ds}),\"—\")")
                cc.number_format = MULT
            r += 1
        r += 1
        ws.cell(row=r, column=1,
                value=f"Table 2 recalculates live. Year-1 DSCR at a {base_rate:.2%} rate assumes the Year-1 payment "
                      f"structure implied by Asmp_IO. Covenant: {eng._get(a,'debt','dscr_min',default=1.25):.2f}x.").font = Font(
            italic=True, size=9, color="808080")

    # ---------------- Checks ----------------

    def sheet_checks(self) -> None:
        ws = self.wb.create_sheet("Checks")
        self.title(ws, "INTEGRITY CHECKS — every row must read OK", 4)
        ws.column_dimensions["A"].width = 52
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 60
        R, h, last = CF, col(self.hold), col(self.hold)

        ws["A3"] = "Check"
        ws["B3"] = "Difference"
        ws["C3"] = "Status"
        ws["D3"] = "Why this matters"
        for c in "ABCD":
            ws[f"{c}3"].font = BOLD
            ws[f"{c}3"].fill = LIGHT

        checks = [
            ("Sources = Uses",
             "=(Asmp_Loan+Ret_Equity)-Ret_Basis",
             "Equity plus debt must fund the entire basis, or the return math starts from a false number."),
            ("NOI = EGI − Total OpEx (Yr 1)",
             f"='Cash Flow'!C{R['noi']}-('Cash Flow'!C{R['egi']}+'Cash Flow'!C{R['opex_total']})",
             "Catches a broken subtotal or a row inserted without updating the sum."),
            ("CFBDS = NOI − Capital (Yr 1)",
             f"='Cash Flow'!C{R['cfbds']}-('Cash Flow'!C{R['noi']}+'Cash Flow'!C{R['capital']})",
             "Capital costs must sit below NOI, not inside it."),
            ("CFADS = CFBDS − Debt Service (Yr 1)",
             f"='Cash Flow'!C{R['cfads']}-('Cash Flow'!C{R['cfbds']}+'Cash Flow'!C{R['ds']})",
             "Confirms debt service is deducted once and only once."),
            ("Debt: End = Begin − Principal (Yr 1)",
             f"='Debt Schedule'!C{DS['end']}-('Debt Schedule'!C{DS['begin']}-'Debt Schedule'!C{DS['principal']})",
             "The amortization schedule must actually reconcile."),
            ("Debt: Service = Interest + Principal (Yr 1)",
             f"='Debt Schedule'!C{DS['service']}-('Debt Schedule'!C{DS['interest']}+'Debt Schedule'!C{DS['principal']})",
             "Splits must sum to the payment."),
            ("Exit value = Forward NOI ÷ Exit Cap",
             f"='Cash Flow'!{h}{R['gross_exit']}-'Cash Flow'!{h}{R['fwd_noi']}/Asmp_ExitCap",
             "Guards the single most common CRE error: capping in-place NOI instead of forward NOI."),
            ("Net equity proceeds reconcile",
             f"='Cash Flow'!{h}{R['net_equity']}-('Cash Flow'!{h}{R['net_sale']}+'Cash Flow'!{h}{R['payoff']})",
             "Sale proceeds less payoff must equal what the equity actually receives."),
            ("Levered total CF = operating + reversion (exit yr)",
             f"='Cash Flow'!{h}{R['lev_total']}-('Cash Flow'!{h}{R['lev_op']}+'Cash Flow'!{h}{R['net_equity']})",
             "The classic double-count: exit proceeds landing in both the operating row and the total row."),
            ("Equity multiple ≥ 1 implies positive total return",
             f"=IF(Ret_Equity>0,(SUM('Cash Flow'!C{R['lev_op']}:{last}{R['lev_op']})"
             f"+'Cash Flow'!{h}{R['net_equity']})/Ret_Equity-1,0)",
             "Informational: negative means the deal returns less than invested capital."),
        ]

        r = 4
        for label, formula, why in checks:
            ws[f"A{r}"] = label
            ws[f"B{r}"] = formula
            ws[f"B{r}"].number_format = '#,##0.00;(#,##0.00);"—"'
            informational = label.startswith("Equity multiple")
            ws[f"C{r}"] = (f'=IF(ABS(B{r})<0.01,"OK","CHECK")' if not informational
                           else f'=IF(B{r}>=0,"OK","REVIEW")')
            ws[f"C{r}"].alignment = Alignment(horizontal="center")
            ws[f"C{r}"].font = BOLD
            ws[f"D{r}"] = why
            ws[f"D{r}"].font = Font(size=9, color="808080")
            ws[f"D{r}"].alignment = Alignment(wrap_text=True, vertical="top")
            r += 1

        r += 1
        ws[f"A{r}"] = "Any row reading CHECK means a formula was edited in a way that broke the model's internal logic. Fix it before circulating."
        ws[f"A{r}"].font = Font(italic=True, size=9, color="C00000")

    # ---------------- README ----------------

    def sheet_readme(self) -> None:
        ws = self.wb.create_sheet("README")
        self.title(ws, "HOW TO USE THIS MODEL", 3)
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 110
        a = self.a
        deal = a.get("deal", {})

        lines: list[tuple[str, str]] = [
            ("h", "WHAT THIS IS"),
            ("p", f"A formula-driven acquisition underwriting model for "
                  f"{deal.get('name', 'this deal')} ({deal.get('property_type', 'n/a')}). "
                  f"{self.hold}-year hold. Every projected cell is a live formula — nothing is typed in twice."),
            ("h", "COLOR CONVENTION (industry standard)"),
            ("p", "BLUE text  = an input. These are the only cells you should type into."),
            ("p", "BLACK text = a formula calculated on the same sheet."),
            ("p", "GREEN text = a link pulling from another sheet."),
            ("p", "AMBER fill = an assumption with no documented source. Fill in the Source column."),
            ("h", "HOW TO FLEX IT"),
            ("p", "Change any blue cell on the Assumptions tab. Cash Flow, Debt Schedule, Returns, and Table 2 "
                  "of the Sensitivity tab all recalculate automatically."),
            ("p", "Table 1 of the Sensitivity tab holds computed VALUES, not formulas, because IRR requires "
                  "re-running the entire projection. Two options: re-run the builder, or convert it to a native "
                  "Excel Data Table."),
            ("h", "TO MAKE TABLE 1 LIVE IN EXCEL"),
            ("p", "1. Put =Returns!<unlevered IRR cell> in the top-left corner cell of the grid."),
            ("p", "2. Select the whole grid including the row and column headers."),
            ("p", "3. Data ▸ What-If Analysis ▸ Data Table."),
            ("p", "4. Row input cell = Asmp_RentGrowth.  Column input cell = Asmp_ExitCap."),
            ("p", "This only works in desktop Excel. LibreOffice and Google Sheets do not round-trip "
                  "two-variable data tables."),
            ("h", "CHECK THE CHECKS TAB FIRST"),
            ("p", "Every row on the Checks tab must read OK. A CHECK status means an edit broke the model's "
                  "internal arithmetic. This is the fastest review you can do on any model, including one you "
                  "did not build."),
            ("h", "CONVENTIONS THIS MODEL FOLLOWS"),
            ("p", "• Exit value divides the exit cap into FORWARD NOI (the year after sale), not the final "
                  "operating year. That is why the pro forma projects one extra year."),
            ("p", "• NOI is before capital reserve, TI, LC, and debt service."),
            ("p", "• Debt yield is reported on the ORIGINAL loan amount (the lender covenant). A current-balance "
                  "row is shown separately for surveillance and is not the covenant."),
            ("p", "• Vacancy and credit loss apply to Potential Gross Income."),
            ("p", "• All rates are decimal fractions: 6.5% is stored as 0.065."),
        ]

        if self.warnings:
            lines.append(("h", "⚠ WARNINGS RAISED WHEN THIS MODEL WAS BUILT"))
            lines.extend(("w", w) for w in self.warnings)

        r = 3
        for kind, text in lines:
            if kind == "h":
                self.section(ws, r, text, 3)
            else:
                c = ws.cell(row=r, column=2, value=text)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.font = Font(size=10, color="C00000" if kind == "w" else "000000")
                ws.row_dimensions[r].height = max(15, 14 * (len(text) // 105 + 1))
            r += 1


def main() -> int:
    if len(sys.argv) < 3:
        print(__doc__)
        return 2
    assumptions = eng.load(sys.argv[1])
    builder = ModelBuilder(assumptions)
    wb = builder.build()
    wb.save(sys.argv[2])

    out = {
        "status": "built",
        "output": sys.argv[2],
        "sheets": wb.sheetnames,
        "warnings": builder.warnings,
        "headline": {
            k: builder.m["returns"][k]
            for k in ("going_in_cap_rate", "unlevered_irr", "levered_irr",
                      "equity_multiple", "equity_required", "min_dscr")
        },
    }
    print(json.dumps(out, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
