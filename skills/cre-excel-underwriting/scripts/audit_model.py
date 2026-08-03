"""
Audit ANY CRE spreadsheet — one you built, or one a broker or seller sent you.

    python3 audit_model.py model.xlsx [--json]

This does not need the model to have been produced by build_model.py. It reads
the workbook structurally and reports the defects that actually mislead people:
formula errors, assumptions buried inside formulas, typed-over "plugs" in a
calculated row, links to files you do not have, and percentage-unit mistakes.

Findings are ranked HIGH / MEDIUM / LOW. Nothing here is a substitute for
reading the model, but it finds in seconds what usually takes an hour.
"""

from __future__ import annotations

import json
import re
import sys
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ERROR_SENTINELS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!", "Err:")

# Integers that are structural rather than assumptions: months, quarters, days,
# percent conversion, and the small counts used in row/column arithmetic.
BENIGN_NUMBERS = {0, 1, 2, 3, 4, 6, 12, 24, 30, 36, 52, 60, 100, 360, 365, 1000, 1000000}

CAPLIKE = re.compile(r"cap|yield|rate|growth|vacancy|ltv|margin|escalat|discount", re.I)
STRING_LITERAL = re.compile(r'"[^"]*"')
CELL_REF = re.compile(r"\$?[A-Z]{1,3}\$?\d{1,7}")
NUMBER = re.compile(r"(?<![A-Za-z0-9_.$])(\d+\.?\d*)(?![A-Za-z0-9_(])")


class Finding(dict):
    def __init__(self, severity, kind, location, detail, why):
        super().__init__(severity=severity, kind=kind, location=location, detail=detail, why=why)


def strip_noise(formula: str) -> str:
    """Remove string literals and cell references so only real constants remain."""
    f = STRING_LITERAL.sub("", formula)
    f = re.sub(r"'[^']*'!", "", f)          # 'Sheet Name'!
    f = re.sub(r"\b[A-Za-z_][A-Za-z0-9_.]*\b(?=\()", "", f)  # function names
    f = CELL_REF.sub("", f)
    return f


def audit(path: str) -> dict:
    wb = load_workbook(path, data_only=False)
    try:
        wbv = load_workbook(path, data_only=True)
    except Exception:
        wbv = None

    findings: list[Finding] = []
    stats = {"sheets": len(wb.worksheets), "formulas": 0, "constants": 0, "blank": 0}
    per_sheet = defaultdict(lambda: {"formulas": 0, "constants": 0})

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if v is None:
                    stats["blank"] += 1
                elif isinstance(v, str) and v.startswith("="):
                    stats["formulas"] += 1
                    per_sheet[ws.title]["formulas"] += 1
                else:
                    stats["constants"] += 1
                    per_sheet[ws.title]["constants"] += 1

    # ---- 1. Error cells (HIGH) --------------------------------------------
    if wbv:
        for ws in wbv.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and any(
                        c.value.startswith(e) for e in ERROR_SENTINELS
                    ):
                        findings.append(Finding(
                            "HIGH", "formula_error", f"{ws.title}!{c.coordinate}", c.value,
                            "An error cell propagates into every total that depends on it. "
                            "No model with a live error should be circulated.",
                        ))

    # ---- 2. External workbook links (HIGH) --------------------------------
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("=") and "[" in c.value:
                    findings.append(Finding(
                        "HIGH", "external_link", f"{ws.title}!{c.coordinate}", c.value[:90],
                        "This pulls from another workbook. If you do not have that file, the "
                        "value you are looking at is a stale cached number, not a live one.",
                    ))

    # ---- 3. Assumptions hardcoded inside formulas (MEDIUM/HIGH) -----------
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                # IRR/XIRR/RATE take a seed guess, and integrity checks compare
                # against a tolerance. Both are structural constants, not
                # assumptions — flagging them trains people to ignore the tool.
                if re.search(r"\b(IRR|XIRR|RATE)\s*\(", v, re.I):
                    continue
                if re.search(r"[<>]=?\s*0?\.0*[015]\b", v):
                    continue

                cleaned = strip_noise(v)
                for raw in NUMBER.findall(cleaned):
                    try:
                        num = float(raw)
                    except ValueError:
                        continue
                    if num in BENIGN_NUMBERS or num == int(num) and abs(num) < 5:
                        continue
                    # A bare multiplier like 1.03 or 1.08 is a growth rate written
                    # as a factor. It is the most common way an assumption hides
                    # inside a formula, so catch it alongside the 0.xx form.
                    if 1 < num < 2 and num != int(num):
                        findings.append(Finding(
                            "MEDIUM", "hardcoded_rate", f"{ws.title}!{c.coordinate}",
                            f"growth factor {num} in: {v[:70]}",
                            "A growth rate written as a multiplier (1.03 = 3%) is an assumption "
                            "buried in a formula. Move it to the assumptions tab and reference it.",
                        ))
                    # A decimal between 0 and 1 inside a formula is almost always
                    # a rate that belongs on the assumptions sheet.
                    elif 0 < num < 1:
                        findings.append(Finding(
                            "MEDIUM", "hardcoded_rate", f"{ws.title}!{c.coordinate}",
                            f"{num} in: {v[:70]}",
                            "A rate typed into a formula cannot be flexed, cannot be sensitised, "
                            "and will not show up when someone reviews the assumptions tab.",
                        ))
                    elif num >= 1000:
                        findings.append(Finding(
                            "MEDIUM", "hardcoded_amount", f"{ws.title}!{c.coordinate}",
                            f"{num:,.0f} in: {v[:70]}",
                            "A dollar amount typed into a formula is invisible to review and "
                            "silently survives every scenario change.",
                        ))

    # ---- 4. Plugs: typed numbers interrupting a calculated row (HIGH) -----
    # A typed value BEFORE the first formula is an input feeding the row, which
    # is exactly how a well-built model starts a projection. A typed value AFTER
    # the row has started calculating is the "someone overwrote a formula to make
    # it tie" pattern. Only the second is a defect, so anchor on that.
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            cells = [c for c in row if c.value is not None and c.column > 1]
            if len(cells) < 4:
                continue
            formulas = [c for c in cells if isinstance(c.value, str) and c.value.startswith("=")]
            if len(formulas) < 3:
                continue
            first_formula_col = min(c.column for c in formulas)
            intruders = [
                c for c in cells
                if isinstance(c.value, (int, float)) and c.column > first_formula_col
            ]
            if intruders:
                label = ws.cell(row=cells[0].row, column=1).value
                findings.append(Finding(
                    "HIGH", "possible_plug", f"{ws.title}!row {cells[0].row}",
                    f"{len(formulas)} formulas, typed values at "
                    f"{', '.join(c.coordinate for c in intruders[:4])}"
                    + (f"  (row label: {str(label)[:40]})" if label else ""),
                    "A hardcoded value sitting inside an otherwise calculated row usually means "
                    "someone typed over a formula to force a number. It will not update.",
                ))

    # ---- 5. Percentage-unit mistakes (HIGH) -------------------------------
    if wbv:
        for ws in wbv.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if not isinstance(c.value, (int, float)):
                        continue
                    label = ws.cell(row=c.row, column=1).value
                    if not (isinstance(label, str) and CAPLIKE.search(label)):
                        continue
                    fmt = wb[ws.title][c.coordinate].number_format or ""
                    if "%" in fmt:
                        continue  # displayed as a percent, so the stored decimal is right
                    if 1.5 < c.value <= 100:
                        findings.append(Finding(
                            "HIGH", "percent_unit", f"{ws.title}!{c.coordinate}",
                            f"{label[:45]} = {c.value} (not percent-formatted)",
                            "A rate stored as 6.5 instead of 0.065 computes cleanly and is wrong "
                            "by 100x. This is the single most expensive unit error in CRE models.",
                        ))

    # ---- 6. Sheets that are entirely typed values (LOW) -------------------
    # An assumptions/input/cover tab is supposed to be all typed values — that
    # is the whole point of separating inputs from calculations.
    input_tab = re.compile(r"assum|input|cover|readme|read me|notes?|source|legend|data", re.I)
    for name, s in per_sheet.items():
        if input_tab.search(name):
            continue
        if s["formulas"] == 0 and s["constants"] > 25:
            findings.append(Finding(
                "LOW", "static_sheet", name,
                f"{s['constants']} values, 0 formulas",
                "A tab with no formulas is a paste of someone else's output. Fine as a source "
                "document; dangerous if the model is supposed to be calculating it.",
            ))

    order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    findings.sort(key=lambda f: (order[f["severity"]], f["kind"]))

    counts = defaultdict(int)
    for f in findings:
        counts[f["severity"]] += 1

    return {
        "file": path,
        "sheets": wb.sheetnames,
        "stats": stats,
        "summary": {s: counts.get(s, 0) for s in ("HIGH", "MEDIUM", "LOW")},
        "findings": findings,
    }


def render(report: dict) -> str:
    out = [f"CRE MODEL AUDIT — {report['file']}", "=" * 70]
    s = report["stats"]
    out.append(
        f"{s['sheets']} sheets | {s['formulas']:,} formulas | {s['constants']:,} typed values"
    )
    sm = report["summary"]
    out.append(f"HIGH: {sm['HIGH']}   MEDIUM: {sm['MEDIUM']}   LOW: {sm['LOW']}")
    out.append("")

    if not report["findings"]:
        out.append("No structural defects found. This checks mechanics, not assumptions —")
        out.append("the numbers can still be wrong even when the wiring is clean.")
        return "\n".join(out)

    grouped: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for f in report["findings"]:
        grouped[(f["severity"], f["kind"])].append(f)

    for (sev, kind), items in grouped.items():
        out.append(f"[{sev}] {kind.replace('_', ' ').upper()}  ({len(items)})")
        out.append(f"  Why it matters: {items[0]['why']}")
        for f in items[:10]:
            out.append(f"    • {f['location']}: {f['detail']}")
        if len(items) > 10:
            out.append(f"    … and {len(items) - 10} more")
        out.append("")
    return "\n".join(out)


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 2
    report = audit(sys.argv[1])
    if "--json" in sys.argv:
        print(json.dumps(report, indent=2, default=str))
    else:
        print(render(report))
    return 1 if report["summary"]["HIGH"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
