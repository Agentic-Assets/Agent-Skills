"""
Recalculate a built workbook with LibreOffice and prove its formulas agree with
the Python engine.

    python3 verify_model.py assumptions.json model.xlsx

This is the step that turns "the spreadsheet looks right" into "the spreadsheet
computes the same numbers as an independent implementation of the documented
math". Run it every time you build a model, and after any hand edit to formulas.

Exit code 0 = every checked value agrees and no Excel error cells exist.
"""

from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import tempfile

from openpyxl import load_workbook

import cre_engine as eng
from build_model import CF, DS, col

ERRORS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!", "Err:")
TOL_ABS = 1.00      # dollars
TOL_REL = 1e-6      # for rates and multiples


def macro_dir() -> str:
    """
    Locate the LibreOffice user profile that the running soffice will read.

    The macro has to land inside that profile or it is never found, and the
    recalculation silently does nothing — which reads as "no calculated values"
    rather than as an error. The path differs by platform.
    """
    if sys.platform == "darwin":
        return os.path.expanduser(
            "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
        )
    if os.name == "nt":
        return os.path.join(
            os.environ.get("APPDATA", os.path.expanduser("~")),
            "LibreOffice", "4", "user", "basic", "Standard",
        )
    return os.path.expanduser("~/.config/libreoffice/4/user/basic/Standard")


def recalc(path: str, timeout: int = 180) -> str:
    """
    Hard-recalculate via LibreOffice and return the path to the recalculated copy.

    LibreOffice will not recalculate an .xlsx on load by default, so we drive it
    with a Basic macro that calls calculateAll() and saves. Converting without
    the macro silently returns the original (empty) cached values, which would
    make verification pass against nothing.
    """
    if not shutil.which("soffice"):
        raise RuntimeError(
            "LibreOffice ('soffice') not found. Install it to verify formulas:\n"
            "  Debian/Ubuntu: apt-get install libreoffice-calc\n"
            "  macOS:         brew install --cask libreoffice\n"
            "Without it the workbook is still correct by construction, but you lose "
            "the independent Excel-vs-engine proof."
        )

    src = os.path.abspath(path)
    workdir = tempfile.mkdtemp(prefix="crecalc-")
    target = os.path.join(workdir, os.path.basename(src))
    shutil.copy(src, target)

    macros = macro_dir()
    os.makedirs(macros, exist_ok=True)
    with open(os.path.join(macros, "Recalc.xba"), "w") as f:
        f.write(
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" '
            '"module.dtd">\n'
            '<script:module xmlns:script="http://openoffice.org/2000/script" '
            'script:name="Recalc" script:language="StarBasic">'
            "Sub HardRecalc(sUrl as String)\n"
            '  Dim oDoc, args(0) as new com.sun.star.beans.PropertyValue\n'
            '  args(0).Name = "Hidden" : args(0).Value = True\n'
            "  oDoc = StarDesktop.loadComponentFromUrl(sUrl, \"_blank\", 0, args())\n"
            "  oDoc.calculateAll()\n"
            "  oDoc.store()\n"
            "  oDoc.close(False)\n"
            "End Sub\n"
            "</script:module>\n"
        )
    with open(os.path.join(macros, "script.xlb"), "w") as f:
        f.write(
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" '
            '"library.dtd">\n'
            '<library:library xmlns:library="http://openoffice.org/2000/library" '
            'library:name="Standard" library:readonly="false" library:passwordprotected="false">'
            '<library:element library:name="Recalc"/></library:library>\n'
        )

    url = "file://" + target.replace(" ", "%20")
    subprocess.run(
        [
            "soffice", "--headless", "--norestore", "--nolockcheck", "--nodefault",
            f'macro:///Standard.Recalc.HardRecalc("{url}")',
        ],
        capture_output=True,
        timeout=timeout,
        check=False,
    )
    return target


def scan_errors(wb) -> list[str]:
    found = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and any(c.value.startswith(e) for e in ERRORS):
                    found.append(f"{ws.title}!{c.coordinate} = {c.value}")
    return found


PLACEHOLDERS = ("—", "-", "", "n/a", "N/A", "n/a — all cash")


def close(a, b, tol_abs=TOL_ABS, tol_rel=TOL_REL) -> bool:
    # A formula that legitimately has no answer renders a placeholder string in
    # Excel and None in the engine. Treat those as agreeing rather than as a
    # mismatch, or every all-cash deal reports a phantom DSCR failure.
    if isinstance(a, str) and a.strip() in PLACEHOLDERS:
        a = None
    if a is None or b is None:
        return a is None and b is None
    if not isinstance(a, (int, float)) or not isinstance(b, (int, float)):
        return False
    if abs(b) < 10:  # rates, multiples
        return abs(a - b) <= max(tol_rel, abs(b) * tol_rel)
    return abs(a - b) <= tol_abs


def verify(assumptions_path: str, xlsx_path: str) -> dict:
    a = eng.load(assumptions_path)
    m = eng.compute(a)
    hold = m["hold_years"]
    n = hold + 1

    recalced = recalc(xlsx_path)
    wb = load_workbook(recalced, data_only=True)

    errors = scan_errors(wb)
    cf, dsh, ret = wb["Cash Flow"], wb["Debt Schedule"], wb["Returns"]

    def rv(label: str):
        for r in range(3, ret.max_row + 1):
            if ret.cell(row=r, column=1).value == label:
                return ret.cell(row=r, column=2).value
        return None

    comparisons: list[dict] = []

    def cmp(name, excel, python_val, tol_abs=TOL_ABS):
        ok = close(excel, python_val, tol_abs)
        comparisons.append({
            "check": name,
            "excel": excel,
            "engine": python_val,
            "match": ok,
            "delta": (excel - python_val) if isinstance(excel, (int, float))
                     and isinstance(python_val, (int, float)) else None,
        })

    rows = m["rows"]
    for y in range(1, n + 1):
        c = col(y)
        cmp(f"Y{y} EGI", cf[f"{c}{CF['egi']}"].value, rows[y - 1]["egi"])
        cmp(f"Y{y} NOI", cf[f"{c}{CF['noi']}"].value, rows[y - 1]["noi"])
        cmp(f"Y{y} CFBDS", cf[f"{c}{CF['cfbds']}"].value, rows[y - 1]["cfbds"])
        cmp(f"Y{y} Debt service", dsh[f"{c}{DS['service']}"].value, rows[y - 1]["debt_service"])
        cmp(f"Y{y} Debt ending balance", dsh[f"{c}{DS['end']}"].value, rows[y - 1]["debt_end"])
        cmp(f"Y{y} CFADS", cf[f"{c}{CF['cfads']}"].value, rows[y - 1]["cfads"])

    h = col(hold)
    ex = m["exit"]
    cmp("Forward NOI", cf[f"{h}{CF['fwd_noi']}"].value, ex["forward_noi"])
    cmp("Gross exit value", cf[f"{h}{CF['gross_exit']}"].value, ex["gross_exit_value"])
    cmp("Net sale proceeds", cf[f"{h}{CF['net_sale']}"].value, ex["net_sale_proceeds"])
    cmp("Net equity proceeds", cf[f"{h}{CF['net_equity']}"].value, ex["net_equity_proceeds"])

    r = m["returns"]
    cmp("Equity required", rv("Equity required"), r["equity_required"])
    cmp("Going-in cap rate", rv("Going-in cap rate"), r["going_in_cap_rate"])
    cmp("Unlevered IRR", rv("Unlevered IRR"), r["unlevered_irr"])
    cmp("Levered IRR", rv("Levered IRR"), r["levered_irr"])
    cmp("Equity multiple", rv("Equity multiple"), r["equity_multiple"])
    cmp("Year 1 cash-on-cash", rv("Year 1 cash-on-cash"), r["year1_cash_on_cash"])
    cmp("DCF indicated value", rv("DCF indicated value"), r["dcf_indicated_value"])
    cmp("NPV vs. total basis", rv("NPV vs. total basis"), r["npv_vs_basis"])
    cmp("Minimum DSCR over hold", rv("Minimum DSCR over hold"), r["min_dscr"])

    # Integrity checks must all read OK.
    chk = wb["Checks"]
    failed_checks = []
    for row in range(4, chk.max_row + 1):
        status = chk.cell(row=row, column=3).value
        if status in ("CHECK", "REVIEW"):
            failed_checks.append(f"{chk.cell(row=row, column=1).value} -> {status}")

    mismatches = [c for c in comparisons if not c["match"]]

    # If EVERY Excel value came back empty, the recalculation never happened —
    # report that plainly instead of drowning the user in phantom mismatches.
    # The usual cause is LibreOffice installed without the Calc filter
    # (libreoffice-core alone cannot open .xlsx).
    if comparisons and all(c["excel"] is None for c in comparisons):
        return {
            "status": "fail",
            "compared": len(comparisons),
            "mismatches": [],
            "excel_error_cells": [],
            "failed_integrity_checks": [],
            "fatal": (
                "LibreOffice produced no calculated values — the recalculation did not run. "
                "Most often this means the Calc filter is missing: `apt-get install libreoffice-calc` "
                "(libreoffice-core by itself cannot open .xlsx). Also make sure no other soffice "
                "process is holding the profile lock."
            ),
        }

    return {
        "status": "pass" if not mismatches and not errors and not failed_checks else "fail",
        "compared": len(comparisons),
        "mismatches": mismatches,
        "excel_error_cells": errors,
        "failed_integrity_checks": failed_checks,
    }


def main() -> int:
    if len(sys.argv) < 3:
        print(__doc__)
        return 2
    result = verify(sys.argv[1], sys.argv[2])
    print(json.dumps(result, indent=2, default=str))
    return 0 if result["status"] == "pass" else 1


if __name__ == "__main__":
    raise SystemExit(main())
